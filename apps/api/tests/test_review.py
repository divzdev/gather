"""Review: scoping, aggregation, blind rounds, assignment.

The aggregate rules are the point. What is *excluded* from a submission's score —
pending reviews, conflicts of interest, free-text criteria, AI scores — is what
decides whether the number an organizer sorts by means anything.
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal

import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.core.tenancy import tenancy_disabled
from app.models import (
    CriterionKind,
    Event,
    EventStatus,
    Form,
    FormKind,
    Organization,
    OrgMember,
    Role,
    Speaker,
    Submission,
    SubmissionSpeaker,
    SubmissionStatus,
    User,
)

PASSWORD = "correct horse battery staple"
FORM_SCHEMA = {
    "sections": [
        {
            "key": "s",
            "title": "Proposal",
            "fields": [
                {"key": "abstract", "type": "long_text", "label": "Abstract"},
                {
                    "key": "speaker_bio",
                    "type": "long_text",
                    "label": "Bio",
                    "identity_bearing": True,
                },
            ],
        }
    ],
    "logic": [],
}


@dataclass
class World:
    headers: dict[str, str]
    reviewer_headers: dict[str, str]
    reviewer2_headers: dict[str, str]
    event: Event
    reviewer_id: uuid.UUID
    reviewer2_id: uuid.UUID
    submissions: list[uuid.UUID]


async def _login(client: AsyncClient, email: str) -> dict[str, str]:
    response = await client.post("/v1/auth/login", json={"email": email, "password": PASSWORD})
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


@pytest.fixture
async def world(client: AsyncClient, session: AsyncSession) -> World:
    suffix = uuid.uuid4().hex[:8]
    with tenancy_disabled():
        org = Organization(name=f"Org {suffix}", slug=f"org-{suffix}")
        session.add(org)
        await session.flush()
        event = Event(
            org_id=org.id,
            name="DevFlow Conf 2027",
            slug=f"devflow-{suffix}",
            timezone="UTC",
            starts_on=datetime(2027, 5, 12).date(),
            ends_on=datetime(2027, 5, 14).date(),
            status=EventStatus.IN_REVIEW,
            cfp_closes_at=datetime.now(UTC) + timedelta(days=1),
        )
        session.add(event)
        await session.flush()
        form = Form(
            org_id=org.id, event_id=event.id, name="CFP", kind=FormKind.CFP, schema=FORM_SCHEMA
        )
        session.add(form)

        people = {}
        for label, role in (
            ("admin", Role.OWNER),
            ("rev1", Role.REVIEWER),
            ("rev2", Role.REVIEWER),
        ):
            user = User(
                email=f"{label}-{suffix}@example.com",
                name=label,
                password_hash=hash_password(PASSWORD),
                email_verified_at=datetime.now(UTC),
            )
            session.add(user)
            await session.flush()
            session.add(OrgMember(org_id=org.id, user_id=user.id, role=role))
            people[label] = user

        speaker = Speaker(org_id=org.id, email=f"spk-{suffix}@example.com", name="Priya Raman")
        session.add(speaker)
        await session.flush()

        ids = []
        for index in range(3):
            submission = Submission(
                org_id=org.id,
                event_id=event.id,
                form_id=form.id,
                code=f"AA{index:04d}"[:6],
                title=f"Proposal {index}",
                answers={"abstract": "About builds.", "speaker_bio": "Priya works at Latticework."},
                status=SubmissionStatus.SUBMITTED,
            )
            session.add(submission)
            await session.flush()
            session.add(
                SubmissionSpeaker(
                    org_id=org.id,
                    event_id=event.id,
                    submission_id=submission.id,
                    speaker_id=speaker.id,
                    is_primary=True,
                )
            )
            ids.append(submission.id)
        await session.commit()

    return World(
        headers=await _login(client, people["admin"].email),
        reviewer_headers=await _login(client, people["rev1"].email),
        reviewer2_headers=await _login(client, people["rev2"].email),
        event=event,
        reviewer_id=people["rev1"].id,
        reviewer2_id=people["rev2"].id,
        submissions=ids,
    )


async def _round(
    client: AsyncClient, world: World, *, blind: bool = False, name: str = "Round 1"
) -> uuid.UUID:
    created = await client.post(
        f"/v1/events/{world.event.id}/review-rounds",
        json={"name": name, "is_blind": blind},
        headers=world.headers,
    )
    round_id = created.json()["id"]
    await client.patch(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}",
        json={"status": "open"},
        headers=world.headers,
    )
    return uuid.UUID(round_id)


async def _criterion(
    client: AsyncClient,
    world: World,
    round_id: uuid.UUID,
    *,
    label: str,
    kind: str = "rating",
    weight: str = "1.00",
    **extra: object,
) -> uuid.UUID:
    response = await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/criteria",
        json={"label": label, "kind": kind, "weight": weight, **extra},
        headers=world.headers,
    )
    return uuid.UUID(response.json()["id"])


async def test_a_round_cannot_be_edited_to_close_before_it_opens(
    client: AsyncClient, world: World
) -> None:
    """`RoundCreate` has checked this since it was written. The edit path never
    did, so the guard was reachable only by getting it wrong the first time —
    one PATCH afterwards put the round into the state creation refuses.
    """
    created = await client.post(
        f"/v1/events/{world.event.id}/review-rounds",
        json={
            "name": "Screening",
            "opens_at": "2027-05-01T09:00:00Z",
            "closes_at": "2027-05-08T17:00:00Z",
        },
        headers=world.headers,
    )
    assert created.status_code == 201
    round_id = created.json()["id"]

    # One-sided: only the close moves, and it lands before the stored open.
    response = await client.patch(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}",
        json={"closes_at": "2027-04-01T09:00:00Z"},
        headers=world.headers,
    )

    assert response.status_code == 422, response.text
    assert response.json()["error"]["code"] == "VALIDATION_FAILED"


async def test_a_form_cannot_close_before_it_opens(client: AsyncClient, world: World) -> None:
    """Neither half of the form window was ordered, on create or on edit — a
    form could open in June and close in May, which reads to the public page as
    a call that is never open."""
    response = await client.post(
        f"/v1/events/{world.event.id}/forms",
        json={
            "name": "Backwards",
            "kind": "cfp",
            "schema": {"sections": [], "logic": [], "settings": {}},
            "opens_at": "2027-06-01T09:00:00Z",
            "closes_at": "2027-05-01T09:00:00Z",
        },
        headers=world.headers,
    )

    assert response.status_code == 422, response.text
    assert response.json()["error"]["code"] == "VALIDATION_FAILED"


async def test_two_independent_rounds_coexist(client: AsyncClient, world: World) -> None:
    await _round(client, world, name="Screening")
    await _round(client, world, name="Final")

    listed = await client.get(f"/v1/events/{world.event.id}/review-rounds", headers=world.headers)

    assert [r["name"] for r in listed.json()] == ["Screening", "Final"]


async def test_scorecard_supports_rating_select_and_text(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    await _criterion(client, world, round_id, label="Relevance")
    await _criterion(
        client,
        world,
        round_id,
        label="Recommendation",
        kind="select",
        choices=[{"value": 1, "label": "Reject"}, {"value": 5, "label": "Strong accept"}],
    )
    await _criterion(client, world, round_id, label="Notes", kind="text", is_required=False)

    listed = await client.get(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/criteria", headers=world.headers
    )

    assert [c["kind"] for c in listed.json()] == ["rating", "select", "text"]


async def test_a_select_criterion_needs_choices(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    response = await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/criteria",
        json={"label": "Recommendation", "kind": "select"},
        headers=world.headers,
    )

    assert response.status_code == 422


async def test_scale_min_above_max_is_rejected(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    response = await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/criteria",
        json={"label": "Backwards", "scale_min": 9, "scale_max": 2},
        headers=world.headers,
    )

    assert response.status_code == 422


async def test_reviewer_queue_contains_exactly_their_assignments(
    client: AsyncClient, world: World
) -> None:
    round_id = await _round(client, world)
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )

    mine = await client.get(
        f"/v1/events/{world.event.id}/review/queue?round_id={round_id}",
        headers=world.reviewer_headers,
    )
    theirs = await client.get(
        f"/v1/events/{world.event.id}/review/queue?round_id={round_id}",
        headers=world.reviewer2_headers,
    )

    assert [i["submission_id"] for i in mine.json()] == [str(world.submissions[0])]
    assert theirs.json() == []


async def test_reading_an_unassigned_proposal_is_forbidden_not_empty(
    client: AsyncClient, world: World
) -> None:
    """403, not 404 or an empty body — the boundary is deliberate, not accidental."""
    round_id = await _round(client, world)
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )

    response = await client.get(
        f"/v1/events/{world.event.id}/review/submissions/{world.submissions[1]}"
        f"?round_id={round_id}",
        headers=world.reviewer_headers,
    )

    assert response.status_code == 403
    assert response.json()["error"]["code"] == "ROLE_REQUIRED"


async def test_blind_round_strips_identity_from_the_payload(
    client: AsyncClient, world: World
) -> None:
    """Stripped server-side: hiding it in the UI leaves it in the network tab."""
    round_id = await _round(client, world, blind=True)
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )

    response = await client.get(
        f"/v1/events/{world.event.id}/review/submissions/{world.submissions[0]}"
        f"?round_id={round_id}",
        headers=world.reviewer_headers,
    )

    body = response.json()
    assert body["is_blind"] is True
    assert body["speakers"] == []
    assert "speaker_bio" not in body["answers"]
    assert "Priya" not in response.text
    assert body["answers"]["abstract"] == "About builds."


async def test_an_open_round_shows_identity(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world, blind=False)
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )

    response = await client.get(
        f"/v1/events/{world.event.id}/review/submissions/{world.submissions[0]}"
        f"?round_id={round_id}",
        headers=world.reviewer_headers,
    )

    assert [s["name"] for s in response.json()["speakers"]] == ["Priya Raman"]
    assert "speaker_bio" in response.json()["answers"]


async def _score(
    client: AsyncClient,
    world: World,
    round_id: uuid.UUID,
    submission_id: uuid.UUID,
    values: dict[uuid.UUID, object],
    *,
    headers: dict[str, str] | None = None,
    conflict: bool = False,
) -> dict[str, object]:
    response = await client.put(
        f"/v1/events/{world.event.id}/review/submissions/{submission_id}/scores"
        f"?round_id={round_id}",
        json={
            "values": {str(k): v for k, v in values.items()},
            "conflict_of_interest": conflict,
        },
        headers=headers or world.reviewer_headers,
    )
    return dict(response.json())


async def test_weighted_mean_respects_criterion_weights(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    heavy = await _criterion(client, world, round_id, label="Relevance", weight="3.00")
    light = await _criterion(client, world, round_id, label="Novelty", weight="1.00")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )

    result = await _score(client, world, round_id, world.submissions[0], {heavy: 5, light: 1})

    # (5*3 + 1*1) / 4 = 4.00, not the unweighted 3.00
    assert Decimal(result["score_avg"]) == Decimal("4.00")


async def test_free_text_criteria_do_not_enter_the_mean(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    notes = await _criterion(client, world, round_id, label="Notes", kind="text", is_required=False)
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )

    result = await _score(
        client, world, round_id, world.submissions[0], {rating: 4, notes: "Nicely argued"}
    )

    assert Decimal(result["score_avg"]) == Decimal("4.00")


async def test_a_conflict_of_interest_is_excluded_from_the_mean(
    client: AsyncClient, world: World
) -> None:
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={
            "submission_ids": [str(world.submissions[0])],
            "user_ids": [str(world.reviewer_id), str(world.reviewer2_id)],
        },
        headers=world.headers,
    )

    await _score(client, world, round_id, world.submissions[0], {rating: 5})
    result = await _score(
        client,
        world,
        round_id,
        world.submissions[0],
        {rating: 1},
        headers=world.reviewer2_headers,
        conflict=True,
    )

    # The conflicted 1 is ignored; only the honest 5 counts.
    assert Decimal(result["score_avg"]) == Decimal("5.00")


async def test_a_second_reviewer_moves_the_mean(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={
            "submission_ids": [str(world.submissions[0])],
            "user_ids": [str(world.reviewer_id), str(world.reviewer2_id)],
        },
        headers=world.headers,
    )

    await _score(client, world, round_id, world.submissions[0], {rating: 5})
    result = await _score(
        client,
        world,
        round_id,
        world.submissions[0],
        {rating: 3},
        headers=world.reviewer2_headers,
    )

    assert Decimal(result["score_avg"]) == Decimal("4.00")


async def test_rescoring_updates_rather_than_duplicates(client: AsyncClient, world: World) -> None:
    """Save-on-selection depends on this being idempotent."""
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )

    await _score(client, world, round_id, world.submissions[0], {rating: 2})
    result = await _score(client, world, round_id, world.submissions[0], {rating: 5})

    assert Decimal(result["score_avg"]) == Decimal("5.00")


async def test_a_score_outside_the_scale_is_rejected(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )

    response = await client.put(
        f"/v1/events/{world.event.id}/review/submissions/{world.submissions[0]}/scores"
        f"?round_id={round_id}",
        json={"values": {str(rating): 99}},
        headers=world.reviewer_headers,
    )

    assert response.status_code == 400


async def test_changing_a_weight_recomputes_without_losing_scores(
    client: AsyncClient, world: World
) -> None:
    """Editing a live rubric must never destroy what reviewers already entered."""
    round_id = await _round(client, world)
    heavy = await _criterion(client, world, round_id, label="Relevance", weight="1.00")
    light = await _criterion(client, world, round_id, label="Novelty", weight="1.00")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )
    before = await _score(client, world, round_id, world.submissions[0], {heavy: 5, light: 1})
    assert Decimal(before["score_avg"]) == Decimal("3.00")

    await client.patch(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/criteria/{heavy}",
        json={"weight": "3.00"},
        headers=world.headers,
    )

    listed = await client.get(f"/v1/events/{world.event.id}/submissions", headers=world.headers)
    row = next(r for r in listed.json()["data"] if r["id"] == str(world.submissions[0]))
    assert Decimal(row["score_avg"]) == Decimal("4.00")
    assert row["review_count"] == 1


async def test_auto_distribution_respects_a_cap(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    response = await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/auto-distribute",
        json={
            "user_ids": [str(world.reviewer_id), str(world.reviewer2_id)],
            "per_submission": 1,
            "cap_per_reviewer": 1,
        },
        headers=world.headers,
    )

    body = response.json()
    assert body["created"] == 2
    # Three submissions, two reviewers, cap of one each — one goes unassigned and
    # that is reported rather than silently dropped.
    assert body["under_assigned"] == 1


async def test_auto_distribution_spreads_evenly(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/auto-distribute",
        json={"user_ids": [str(world.reviewer_id), str(world.reviewer2_id)], "per_submission": 1},
        headers=world.headers,
    )

    progress = await client.get(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/progress", headers=world.headers
    )
    counts = sorted(r["assigned"] for r in progress.json())

    assert sum(counts) == 3
    assert counts[-1] - counts[0] <= 1


async def test_asking_for_more_reviewers_than_exist_is_rejected(
    client: AsyncClient, world: World
) -> None:
    round_id = await _round(client, world)
    response = await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/auto-distribute",
        json={"user_ids": [str(world.reviewer_id)], "per_submission": 3},
        headers=world.headers,
    )

    assert response.status_code == 400


async def test_progress_reports_assigned_and_completed(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={
            "submission_ids": [str(s) for s in world.submissions[:2]],
            "user_ids": [str(world.reviewer_id)],
        },
        headers=world.headers,
    )
    await _score(client, world, round_id, world.submissions[0], {rating: 4})

    progress = await client.get(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/progress", headers=world.headers
    )

    row = next(r for r in progress.json() if r["user_id"] == str(world.reviewer_id))
    assert (row["assigned"], row["completed"]) == (2, 1)


async def test_nudge_skips_reviewers_who_are_done(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={
            "submission_ids": [str(world.submissions[0])],
            "user_ids": [str(world.reviewer_id), str(world.reviewer2_id)],
        },
        headers=world.headers,
    )
    await _score(client, world, round_id, world.submissions[0], {rating: 4})

    response = await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/nudge", headers=world.headers
    )

    assert response.json() == {"sent": 1, "skipped": 1}


async def test_results_export_is_csv_with_one_row_per_submission(
    client: AsyncClient, world: World
) -> None:
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )
    await _score(client, world, round_id, world.submissions[0], {rating: 4})

    response = await client.get(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/results.csv",
        headers=world.headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    lines = response.text.strip().splitlines()
    assert lines[0].startswith("code,title,speakers,status,reviews,average_score")
    assert len(lines) == 4  # header plus three submissions
    assert "4.00" in response.text
    assert "Priya Raman" in response.text


async def test_a_reviewer_cannot_reach_the_admin_surface(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)

    response = await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/auto-distribute",
        json={"user_ids": [str(world.reviewer_id)], "per_submission": 1},
        headers=world.reviewer_headers,
    )

    assert response.status_code == 403


async def test_scoring_a_closed_round_is_refused(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )
    await client.patch(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}",
        json={"status": "closed"},
        headers=world.headers,
    )

    response = await client.put(
        f"/v1/events/{world.event.id}/review/submissions/{world.submissions[0]}/scores"
        f"?round_id={round_id}",
        json={"values": {str(rating): 4}},
        headers=world.reviewer_headers,
    )

    assert response.status_code == 409


async def test_threshold_advancement_moves_only_the_qualifying(
    client: AsyncClient, world: World
) -> None:
    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.patch(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}",
        json={"advance_rule": {"type": "threshold", "min_score": 4}},
        headers=world.headers,
    )
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={
            "submission_ids": [str(s) for s in world.submissions[:2]],
            "user_ids": [str(world.reviewer_id)],
        },
        headers=world.headers,
    )
    await _score(client, world, round_id, world.submissions[0], {rating: 5})
    await _score(client, world, round_id, world.submissions[1], {rating: 2})

    response = await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/advance", headers=world.headers
    )

    assert response.json() == {"advanced": 1}


async def test_a_manual_round_advances_nothing(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)

    response = await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/advance", headers=world.headers
    )

    assert response.json() == {"advanced": 0}


async def test_criterion_kind_is_stored(client: AsyncClient, world: World) -> None:
    round_id = await _round(client, world)
    criterion_id = await _criterion(
        client,
        world,
        round_id,
        label="Recommendation",
        kind=CriterionKind.SELECT.value,
        choices=[{"value": 1, "label": "No"}, {"value": 5, "label": "Yes"}],
    )

    listed = await client.get(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/criteria", headers=world.headers
    )
    row = next(c for c in listed.json() if c["id"] == str(criterion_id))

    assert row["kind"] == "select"
    assert len(row["choices"]) == 2


async def test_results_also_export_as_a_spreadsheet(client: AsyncClient, world: World) -> None:
    """Programme committees pass spreadsheets around, and a score stored as text
    sorts 10 before 9 — so the number column has to be a number."""
    import io

    from openpyxl import load_workbook

    round_id = await _round(client, world)
    rating = await _criterion(client, world, round_id, label="Relevance")
    await client.post(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/assignments",
        json={"submission_ids": [str(world.submissions[0])], "user_ids": [str(world.reviewer_id)]},
        headers=world.headers,
    )
    await _score(client, world, round_id, world.submissions[0], {rating: 4})

    response = await client.get(
        f"/v1/events/{world.event.id}/review-rounds/{round_id}/results.xlsx",
        headers=world.headers,
    )

    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]

    sheet = load_workbook(io.BytesIO(response.content)).active
    assert sheet is not None
    header = [cell.value for cell in sheet[1]]
    assert header[:6] == ["code", "title", "speakers", "status", "reviews", "average_score"]
    assert sheet.max_row == 4  # header plus three submissions

    scores = [sheet.cell(row=row, column=6).value for row in range(2, 5)]
    assert 4 in scores
    # Numeric, not text. A whole number reads back as int, which is still a
    # number Excel will sort and average; a string is what breaks it.
    assert all(value is None or isinstance(value, int | float) for value in scores)
    assert not any(isinstance(value, str) for value in scores)
